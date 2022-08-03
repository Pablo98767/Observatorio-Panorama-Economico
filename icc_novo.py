from types import coroutine
from numpy import SHIFT_DIVIDEBYZERO, datetime_data, sign
from numpy.lib.nanfunctions import nanvar
import requests
from bs4 import BeautifulSoup as bs
import pandas as pd
import json
from datetime import datetime
from openpyxl import load_workbook
import util as util
from util import *
from upload import *
from github import Github


path_planilha = 'G:/IEL/OBSERVATORIO/ETL/planilhas de dados/icc fgv-iel-jully.xlsx'
num_planilha = 0

wb = load_workbook(filename = path_planilha)
sheet_name = wb.sheetnames[num_planilha]
#access specific sheet
#ws = wb["Tables"]
print('- Planilha acessada')                                                                          




print('- Criando série histórica')     

ws = wb[sheet_name]                                                                                 
df = pd.DataFrame(ws.values)                                                                             
df.columns = df.iloc[0]                                                                                 
df = df.iloc[1:]                                                                                   
df = df[['referencia', 'valores']]                                                                     
df['referencia'] =  pd.to_datetime(df['referencia'],  format='%d%b%Y:%H:%M:%S.%f')
df.sort_values('referencia', inplace =True) 
df = df.reset_index(drop=True)
df.dropna(inplace =True)

df['referencia'] = df['referencia'].astype(str)
df['referencia'] = df['referencia'].str[:7]

shifted = df['valores'].shift(1)
df['variacao_mensal'] = ((df['valores'] - shifted) / shifted) * 100


serie_icc = []
now = datetime.now()
ano_atual = now.year
ano_anterior = ano_atual - 1


for i, row in df.iterrows():
  if row[0][:4] == str(ano_anterior) or row[0][:4] == str(ano_atual):
    mes = row[0][:7].replace('-', '')
    mes = pd.to_datetime(mes, format='%Y%m', errors='coerce')
    mes = mes.strftime("%Y-%m-%dT%H:%M:%SZ")
    serie_icc.append({'x': mes, 'y': round(row[2],2)})
    
    
print('- Série criada')

meses_ano = ['Jan', 'Fev', 'Mar', 'Abr', 'Mai', 'Jun', 'Jul', 'Ago', 'Set', 'Out', 'Nov', 'Dez']
ano_referencia = df['referencia'][-1:].iloc[0][:4]
mes_referencia = int(df['referencia'][-1:].iloc[0][-2:])
referencia = meses_ano[mes_referencia-1] + '/' + ano_referencia


valor_periodo_anterior = serie_icc[-2:][0]['y']
valor_periodo_atual = serie_icc[-2:][1]['y']
valor_cartao = valor_periodo_atual


if valor_periodo_atual > 0:
  direcao = 'up'
  cor_valor = 'green'
elif valor_periodo_atual < 0:
  direcao = 'down'
  cor_valor = 'red'
  #remove o sinal de negativo
  valor_cartao = str(valor_cartao)[1:]
else:
  direcao = 'right'
  cor_valor = 'gray'
  

icc = {
    'nome': 'Índice de Confiança do Consumidor',
    'descricao': 'Variação mensal',
    'fonte': 'FGV',
    'stats': [
        {
            'titulo': 'Brasil',
            'valor': str(valor_cartao)+'%',
            'direcao': direcao,
            'cor_valor': cor_valor,
            'desc_serie': 'Variação percentual mensal',
            'serie_tipo': 'data',
            'referencia': referencia,
            'y_label': {
                    'prefixo_sufixo': 'sufixo',
                    'label': '%',
            },

            'serie_labels': {
                'x': 'Data',
                'y': 'Variação'
            },
            'serie': serie_icc,
        },
    ]
}

path_save_json = util.config['path_save_json']['path']
name_json = 'icc'

with open(path_save_json + name_json + '.json', 'w', encoding='utf-8') as f:
    json.dump(icc, f, ensure_ascii=False, indent=4)


print('- JSON Armazenado')

######################################################
#Upload

upload_files_to_github(name_json)
print('JSON enviado para o GitHub')